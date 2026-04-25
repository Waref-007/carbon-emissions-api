# main.py - Production API layer for Carbon Emissions Engine v3
#
# Aligned with engine_production_v3.py / engine.py:
# - Multi-file CSV/XLS/XLSX upload
# - Multi-sheet workbook parsing
# - Canonical and semi-structured client data support
# - Optional custom factor rows
# - Executive Excel report with charts
# - Audit trail, parser diagnostics, factor transparency, confidence scoring
# - Safer validation, request IDs, runtime metadata, and file-level error handling

from __future__ import annotations

import io
import json
import os
import re
import time
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Annotated, Any, Dict, List, Optional, Tuple

import pandas as pd
import requests
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from engine import run_emissions_engine


# ============================================================
# APP CONFIG
# ============================================================

APP_VERSION = "3.0.0"
MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "25"))
MAX_TOTAL_UPLOAD_MB = int(os.getenv("MAX_TOTAL_UPLOAD_MB", "75"))
MAX_EXCEL_SHEETS = int(os.getenv("MAX_EXCEL_SHEETS", "50"))
MAX_PREVIEW_ERRORS = int(os.getenv("MAX_PREVIEW_ERRORS", "50"))
REPORT_FILENAME = os.getenv("REPORT_FILENAME", "carbon_emissions_audit_report.xlsx")

ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}

DEFAULT_ALLOWED_ORIGINS = [
    "https://gscsustainability.co.uk",
    "https://www.gscsustainability.co.uk",
    "http://localhost:3000",
    "http://localhost:5173",
]

ALLOWED_ORIGINS = [
    origin.strip()
    for origin in os.getenv("ALLOWED_ORIGINS", ",".join(DEFAULT_ALLOWED_ORIGINS)).split(",")
    if origin.strip()
]

app = FastAPI(
    title="Carbon Emissions API",
    version=APP_VERSION,
    description="Production API for carbon emissions calculation, audit trail, and reporting.",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================
# ROUTES: HEALTH / DEBUG
# ============================================================

@app.get("/")
def root() -> Dict[str, Any]:
    return {
        "status": "ok",
        "message": "Carbon emissions API is running",
        "app_version": APP_VERSION,
        "generated_at_utc": utc_now(),
    }


@app.head("/")
def root_head() -> None:
    return None


@app.get("/health")
def health() -> Dict[str, Any]:
    return {
        "status": "healthy",
        "app_version": APP_VERSION,
        "generated_at_utc": utc_now(),
    }


@app.get("/debug-routes")
def debug_routes() -> Dict[str, Any]:
    return {
        "routes": [
            "/",
            "/health",
            "/calculate",
            "/upload-calculate",
            "/upload-calculate-download",
        ],
        "app_version": APP_VERSION,
        "allowed_origins": ALLOWED_ORIGINS,
        "max_upload_mb": MAX_UPLOAD_MB,
        "max_total_upload_mb": MAX_TOTAL_UPLOAD_MB,
    }


# ============================================================
# GENERAL HELPERS
# ============================================================

def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def new_request_id() -> str:
    return uuid.uuid4().hex[:16]


def safe_string(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value) and not isinstance(value, str):
            return ""
    except Exception:
        pass
    return str(value).strip()


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        if isinstance(value, str):
            value = (
                value.replace(",", "")
                .replace("£", "")
                .replace("$", "")
                .replace("€", "")
                .strip()
            )
        if value == "":
            return default
        return float(value)
    except Exception:
        return default


def parse_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y", "on"}:
        return True
    if text in {"false", "0", "no", "n", "off"}:
        return False
    return default


def sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "", safe_string(name))
    return cleaned[:31] if cleaned else "Sheet"


def file_extension(filename: str) -> str:
    name = safe_string(filename).lower()
    for ext in ALLOWED_EXTENSIONS:
        if name.endswith(ext):
            return ext
    return ""


def validate_email(email: str) -> bool:
    text = safe_string(email)
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", text))


def normalize_uploaded_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [
        str(col).strip().lower().replace("\n", " ").replace("\r", " ")
        for col in df.columns
    ]
    return df


def dataframe_to_workbook_records(df: pd.DataFrame) -> List[Dict[str, Any]]:
    df = df.copy().where(pd.notnull(df), None)
    records: List[Dict[str, Any]] = []
    for row in df.itertuples(index=False, name=None):
        records.append({f"col_{idx + 1}": value for idx, value in enumerate(row)})
    return records


def make_json_safe(obj: Any) -> Any:
    if isinstance(obj, dict):
        return {str(k): make_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [make_json_safe(v) for v in obj]
    if isinstance(obj, tuple):
        return [make_json_safe(v) for v in obj]
    if isinstance(obj, pd.DataFrame):
        return make_json_safe(obj.to_dict(orient="records"))
    if isinstance(obj, pd.Series):
        return make_json_safe(obj.to_dict())
    if isinstance(obj, pd.Timestamp):
        return obj.isoformat()
    try:
        if pd.isna(obj) and not isinstance(obj, str):
            return None
    except Exception:
        pass
    return obj


# ============================================================
# PARSER OPTIONS / FACTOR ROWS
# ============================================================

DEFAULT_PARSER_OPTIONS: Dict[str, Any] = {
    "include_fleet_fuel": True,
    "include_business_mileage": False,
    "include_business_mileage_when_fuel_present": False,
    "include_ev_charging_submeter": False,
    "fleet_fuel_type": "Diesel",
    "electricity_country": "Electricity: UK",
    "gas_unit": "kWh (Gross CV)",
}


def build_parser_options(form_values: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    options = dict(DEFAULT_PARSER_OPTIONS)
    if not form_values:
        return options

    bool_keys = {
        "include_fleet_fuel",
        "include_business_mileage",
        "include_business_mileage_when_fuel_present",
        "include_ev_charging_submeter",
    }

    for key, value in form_values.items():
        if value is None or value == "":
            continue
        if key in bool_keys:
            options[key] = parse_bool(value, default=bool(options.get(key, False)))
        else:
            options[key] = value

    return options


def parse_factor_rows_json(factor_rows_json: Optional[str]) -> Optional[List[Dict[str, Any]]]:
    if not safe_string(factor_rows_json):
        return None

    try:
        parsed = json.loads(factor_rows_json)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"factor_rows_json must be valid JSON: {type(exc).__name__}: {str(exc)}",
        )

    if parsed is None:
        return None

    if not isinstance(parsed, list):
        raise HTTPException(status_code=400, detail="factor_rows_json must be a JSON list of factor rows.")

    for i, row in enumerate(parsed):
        if not isinstance(row, dict):
            raise HTTPException(status_code=400, detail=f"factor_rows_json row {i + 1} must be an object.")

    return parsed


# ============================================================
# USER / FORM VALIDATION
# ============================================================

def validate_user_inputs(
    privacy_consent: str,
    contact_consent: str,
    full_name: str,
    email: str,
    company_name: str,
    phone_number: str,
) -> None:
    if not parse_bool(privacy_consent):
        raise HTTPException(status_code=400, detail="Privacy consent is required.")

    if not parse_bool(contact_consent):
        raise HTTPException(status_code=400, detail="Contact consent is required.")

    if not safe_string(full_name):
        raise HTTPException(status_code=400, detail="Full name is required.")

    if not safe_string(email):
        raise HTTPException(status_code=400, detail="Email is required.")

    if not validate_email(email):
        raise HTTPException(status_code=400, detail="A valid email address is required.")

    if not safe_string(company_name):
        raise HTTPException(status_code=400, detail="Company name is required.")

    if not safe_string(phone_number):
        raise HTTPException(status_code=400, detail="Phone number is required.")


def build_user_object(full_name: str, email: str, company_name: str, phone_number: str) -> Dict[str, Any]:
    return {
        "full_name": safe_string(full_name),
        "email": safe_string(email),
        "company_name": safe_string(company_name),
        "phone_number": safe_string(phone_number),
    }


# ============================================================
# UPLOAD READING
# ============================================================

def validate_upload_file(filename: str, content: bytes) -> None:
    if not filename:
        raise HTTPException(status_code=400, detail="Uploaded file is missing a filename.")

    ext = file_extension(filename)
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file format for '{filename}'. Please upload CSV, XLS, or XLSX files.",
        )

    size_mb = len(content) / (1024 * 1024)
    if size_mb > MAX_UPLOAD_MB:
        raise HTTPException(
            status_code=400,
            detail=f"'{filename}' is too large ({round(size_mb, 2)} MB). Maximum allowed is {MAX_UPLOAD_MB} MB.",
        )


def extract_file_sources(uploaded_file: UploadFile) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], int]:
    filename = uploaded_file.filename or "uploaded_file"
    content = uploaded_file.file.read()
    validate_upload_file(filename, content)

    ext = file_extension(filename)
    uploaded_sources: List[Dict[str, Any]] = []
    file_level_errors: List[Dict[str, Any]] = []
    total_bytes = len(content)

    try:
        if ext == ".csv":
            df = pd.read_csv(
                BytesIO(content),
                dtype=object,
                keep_default_na=False,
                engine="python",
                on_bad_lines="skip",
            )
            df = normalize_uploaded_columns(df)

            uploaded_sources.append(
                {
                    "source_file": filename,
                    "source_type": "csv",
                    "sheet_name": None,
                    "records": df.where(pd.notnull(df), None).to_dict(orient="records"),
                    "rows": int(len(df)),
                    "columns": [str(c) for c in df.columns],
                    "size_bytes": total_bytes,
                }
            )
            return uploaded_sources, file_level_errors, total_bytes

        excel_file = pd.ExcelFile(BytesIO(content))
        sheet_names = excel_file.sheet_names[:MAX_EXCEL_SHEETS]

        if not sheet_names:
            file_level_errors.append(
                {
                    "source_file": filename,
                    "error": "Workbook contained no readable sheets.",
                    "severity": "high",
                }
            )
            return uploaded_sources, file_level_errors, total_bytes

        for sheet_name in sheet_names:
            try:
                df_sheet = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    header=None,
                    dtype=object,
                )
                records = dataframe_to_workbook_records(df_sheet)
                uploaded_sources.append(
                    {
                        "source_file": filename,
                        "source_type": "excel_sheet",
                        "sheet_name": sheet_name,
                        "records": records,
                        "rows": int(df_sheet.shape[0]),
                        "columns": [f"col_{i + 1}" for i in range(int(df_sheet.shape[1]))],
                        "size_bytes": total_bytes,
                    }
                )
            except Exception as exc:
                file_level_errors.append(
                    {
                        "source_file": filename,
                        "source_sheet": sheet_name,
                        "error": f"Could not read sheet '{sheet_name}': {type(exc).__name__}: {str(exc)}",
                        "severity": "high",
                    }
                )

        if len(excel_file.sheet_names) > MAX_EXCEL_SHEETS:
            file_level_errors.append(
                {
                    "source_file": filename,
                    "error": f"Workbook has more than {MAX_EXCEL_SHEETS} sheets. Only the first {MAX_EXCEL_SHEETS} were processed.",
                    "severity": "medium",
                }
            )

        return uploaded_sources, file_level_errors, total_bytes

    except HTTPException:
        raise
    except Exception as exc:
        file_level_errors.append(
            {
                "source_file": filename,
                "error": f"Unexpected file read error: {type(exc).__name__}: {str(exc)}",
                "severity": "high",
            }
        )
        return uploaded_sources, file_level_errors, total_bytes


def prepare_uploaded_sources(files: List[UploadFile]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], int, int]:
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded.")

    all_sources: List[Dict[str, Any]] = []
    file_level_errors: List[Dict[str, Any]] = []
    input_row_count = 0
    total_bytes = 0

    for uploaded_file in files:
        sources, errors, file_bytes = extract_file_sources(uploaded_file)
        total_bytes += file_bytes
        all_sources.extend(sources)
        file_level_errors.extend(errors)
        input_row_count += sum(int(src.get("rows", 0) or 0) for src in sources)

    total_mb = total_bytes / (1024 * 1024)
    if total_mb > MAX_TOTAL_UPLOAD_MB:
        raise HTTPException(
            status_code=400,
            detail=f"Uploaded files are too large in total ({round(total_mb, 2)} MB). Maximum allowed is {MAX_TOTAL_UPLOAD_MB} MB.",
        )

    if not all_sources and not file_level_errors:
        raise HTTPException(status_code=400, detail="No readable files were uploaded.")

    return all_sources, file_level_errors, input_row_count, total_bytes


# ============================================================
# ENGINE ORCHESTRATION / RESULT COMBINATION
# ============================================================

def run_engine_for_uploaded_sources(
    uploaded_sources: List[Dict[str, Any]],
    parser_options: Dict[str, Any],
    factor_rows: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    source_results: List[Dict[str, Any]] = []

    for source in uploaded_sources:
        source_parser_options = {
            **parser_options,
            "source_file": source.get("source_file"),
            "source_sheet": source.get("sheet_name"),
            "source_workbook": source.get("source_file"),
        }

        request_json: Dict[str, Any] = {
            "activities": source.get("records", []),
            "parser_options": source_parser_options,
        }

        if factor_rows:
            request_json["factor_rows"] = factor_rows

        result = run_emissions_engine(request_json)

        pdx = result.get("parser_diagnostics", {}) or {}
        pdx["source_file"] = source.get("source_file")
        pdx["source_sheet"] = source.get("sheet_name")
        pdx["source_type"] = source.get("source_type")
        pdx["source_rows"] = source.get("rows", 0)
        result["parser_diagnostics"] = pdx

        source_results.append(result)

    return combine_engine_results(source_results)


def _sum_numeric(values: List[Any]) -> float:
    return round(sum(safe_float(v, 0.0) for v in values), 6)


def _merge_count_dicts(items: List[Dict[str, Any]], possible_key_names: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    if not items:
        return []

    possible_key_names = possible_key_names or ["category", "error", "factor_quality", "title", "metric", "label"]
    counts: Dict[str, int] = {}
    output_key = "label"

    for candidate in possible_key_names:
        if candidate in items[0]:
            output_key = candidate
            break

    for item in items:
        key = ""
        for candidate in possible_key_names:
            key = safe_string(item.get(candidate))
            if key:
                break
        if not key:
            continue
        counts[key] = counts.get(key, 0) + int(safe_float(item.get("count", 1), 1))

    return [{output_key: key, "count": value} for key, value in sorted(counts.items(), key=lambda x: (-x[1], x[0]))]


def _combine_issue_groups(issue_groups: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}

    for item in issue_groups:
        title = safe_string(item.get("title"))
        if not title:
            continue

        if title not in grouped:
            grouped[title] = {
                "title": title,
                "count": int(safe_float(item.get("count", 0), 0)),
                "severity": item.get("severity", "low"),
                "guidance": item.get("guidance", ""),
                "examples": list(item.get("examples", []) or [])[:5],
            }
        else:
            grouped[title]["count"] += int(safe_float(item.get("count", 0), 0))
            grouped[title]["examples"] = (grouped[title]["examples"] + list(item.get("examples", []) or []))[:5]
            severity_rank = {"high": 3, "medium": 2, "low": 1}
            if severity_rank.get(item.get("severity", "low"), 0) > severity_rank.get(grouped[title].get("severity", "low"), 0):
                grouped[title]["severity"] = item.get("severity", "low")

    severity_order = {"high": 0, "medium": 1, "low": 2}
    return sorted(
        grouped.values(),
        key=lambda x: (severity_order.get(x.get("severity", "low"), 9), -int(x.get("count", 0)), x.get("title", "")),
    )


def _build_error_summary(errors: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not errors:
        return []
    df = pd.DataFrame(errors)
    if "error" not in df.columns:
        return []
    grouped = df["error"].astype(str).value_counts().reset_index()
    grouped.columns = ["error", "count"]
    return grouped.to_dict(orient="records")


def _build_combined_data_quality(
    results: List[Dict[str, Any]],
    line_items: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
    parser_diagnostics: List[Dict[str, Any]],
) -> Dict[str, Any]:
    totals = {
        "total_rows": 0,
        "successful_rows": 0,
        "errored_rows": 0,
        "mapped_category_rows": 0,
        "inferred_category_rows": 0,
        "exact_factor_rows": 0,
        "estimated_rows": 0,
        "disclosure_only_rows": 0,
        "official_factor_rows": 0,
        "starter_factor_rows": 0,
        "factor_fallback_rows": 0,
        "extracted_activity_count": 0,
    }

    for result in results:
        dq = result.get("data_quality", {}) or {}
        for key in totals:
            totals[key] += int(safe_float(dq.get(key, 0), 0))

        pdx = result.get("parser_diagnostics", {}) or {}
        if pdx and not dq.get("extracted_activity_count"):
            totals["extracted_activity_count"] += int(safe_float(pdx.get("extracted_activity_count", 0), 0))

    if totals["successful_rows"] == 0 and line_items:
        totals["successful_rows"] = len(line_items)

    if totals["errored_rows"] == 0 and errors:
        totals["errored_rows"] = len(errors)

    if totals["total_rows"] == 0:
        totals["total_rows"] = totals["successful_rows"] + totals["errored_rows"]

    coverage = round((totals["successful_rows"] / totals["total_rows"]) * 100, 2) if totals["total_rows"] > 0 else 0.0

    totals["coverage_percent"] = coverage
    totals["valid_rows_percent"] = coverage
    totals["invalid_rows_percent"] = round((totals["errored_rows"] / totals["total_rows"]) * 100, 2) if totals["total_rows"] > 0 else 0.0
    totals["parser_warning_count"] = sum(len(pdx.get("warnings", []) or []) for pdx in parser_diagnostics)
    totals["excluded_section_count"] = sum(len(pdx.get("excluded_sections", []) or []) for pdx in parser_diagnostics)

    return totals


def _build_confidence_score(data_quality: Dict[str, Any]) -> Dict[str, Any]:
    total_rows = int(data_quality.get("total_rows", 0) or 0)
    coverage = float(data_quality.get("coverage_percent", 0.0) or 0.0)

    if total_rows <= 0:
        return {
            "score": 0,
            "label": "Low",
            "coverage_percent": 0.0,
            "notes": ["No activity rows were available for calculation."],
        }

    score = 100.0
    score -= max(0.0, 100.0 - coverage) * 0.65
    score -= (int(data_quality.get("estimated_rows", 0) or 0) / total_rows) * 20.0
    score -= (int(data_quality.get("starter_factor_rows", 0) or 0) / total_rows) * 25.0
    score -= (int(data_quality.get("factor_fallback_rows", 0) or 0) / total_rows) * 10.0
    score -= min(int(data_quality.get("parser_warning_count", 0) or 0) * 3.0, 12.0)
    score -= min(int(data_quality.get("excluded_section_count", 0) or 0) * 1.0, 8.0)
    score = int(round(max(0.0, min(100.0, score)), 0))

    label = "High" if score >= 80 else "Moderate" if score >= 60 else "Low"

    notes: List[str] = []
    notes.append(f"{coverage}% of extracted activity rows were processed successfully.")
    if data_quality.get("starter_factor_rows", 0):
        notes.append(f"{data_quality.get('starter_factor_rows')} row(s) used starter/placeholder factors.")
    if data_quality.get("factor_fallback_rows", 0):
        notes.append(f"{data_quality.get('factor_fallback_rows')} row(s) used fallback factor matching.")
    if data_quality.get("parser_warning_count", 0):
        notes.append(f"{data_quality.get('parser_warning_count')} parser warning(s) should be reviewed.")
    if data_quality.get("excluded_section_count", 0):
        notes.append(f"{data_quality.get('excluded_section_count')} section/item(s) were excluded, often to avoid double counting.")

    return {
        "score": score,
        "label": label,
        "coverage_percent": coverage,
        "notes": notes,
    }


def combine_engine_results(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    line_items: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    factor_transparency: List[Dict[str, Any]] = []
    parser_diagnostics: List[Dict[str, Any]] = []
    parser_warnings: List[Dict[str, Any]] = []
    issue_groups: List[Dict[str, Any]] = []
    unmapped_categories: List[Dict[str, Any]] = []
    factor_quality_summary: List[Dict[str, Any]] = []

    for result in results:
        line_items.extend(result.get("line_items", []) or [])
        errors.extend(result.get("errors", []) or [])
        factor_transparency.extend(result.get("factor_transparency", []) or [])
        issue_groups.extend(result.get("issue_groups", []) or [])
        unmapped_categories.extend(result.get("unmapped_categories", []) or [])
        factor_quality_summary.extend(result.get("factor_quality_summary", []) or [])

        pdx = result.get("parser_diagnostics", {}) or {}
        if pdx:
            parser_diagnostics.append(pdx)
            for warning in pdx.get("warnings", []) or []:
                parser_warnings.append(
                    {
                        "source_file": pdx.get("source_file", ""),
                        "source_sheet": pdx.get("source_sheet", ""),
                        "error": warning,
                        "severity": "medium",
                        "type": "Parser warning",
                    }
                )

    all_errors = parser_warnings + errors
    df_line = pd.DataFrame(line_items)

    if not df_line.empty and "emissions_kgCO2e" in df_line.columns:
        df_line["emissions_kgCO2e"] = pd.to_numeric(df_line["emissions_kgCO2e"], errors="coerce").fillna(0.0)
        df_line["emissions_tCO2e"] = df_line["emissions_kgCO2e"] / 1000.0

        total_kg = round(float(df_line["emissions_kgCO2e"].sum()), 6)
        total_t = round(total_kg / 1000.0, 6)

        if "scope" in df_line.columns:
            df_scope = (
                df_line.groupby("scope", as_index=False)[["emissions_kgCO2e", "emissions_tCO2e"]]
                .sum()
                .sort_values("emissions_kgCO2e", ascending=False)
                .reset_index(drop=True)
            )
            scope_total = float(df_scope["emissions_kgCO2e"].sum()) if not df_scope.empty else 0.0
            df_scope["percent_of_total"] = df_scope["emissions_kgCO2e"].apply(
                lambda x: round((float(x) / scope_total) * 100, 2) if scope_total > 0 else 0.0
            )
        else:
            df_scope = pd.DataFrame(columns=["scope", "emissions_kgCO2e", "emissions_tCO2e", "percent_of_total"])

        if {"scope", "category"}.issubset(df_line.columns):
            df_scope_category = (
                df_line.groupby(["scope", "category"], as_index=False)[["emissions_kgCO2e", "emissions_tCO2e"]]
                .sum()
                .sort_values("emissions_kgCO2e", ascending=False)
                .reset_index(drop=True)
            )
        else:
            df_scope_category = pd.DataFrame(columns=["scope", "category", "emissions_kgCO2e", "emissions_tCO2e"])

        if "category" in df_line.columns:
            df_top_categories = (
                df_line.groupby("category", as_index=False)[["emissions_kgCO2e", "emissions_tCO2e"]]
                .sum()
                .sort_values("emissions_kgCO2e", ascending=False)
                .head(5)
                .reset_index(drop=True)
            )
        else:
            df_top_categories = pd.DataFrame(columns=["category", "emissions_kgCO2e", "emissions_tCO2e"])

        if "site_name" in df_line.columns:
            df_sites = (
                df_line.assign(site_name=df_line["site_name"].fillna("Unspecified site"))
                .groupby("site_name", as_index=False)[["emissions_kgCO2e", "emissions_tCO2e"]]
                .sum()
                .sort_values("emissions_kgCO2e", ascending=False)
                .head(5)
                .reset_index(drop=True)
            )
        else:
            df_sites = pd.DataFrame(columns=["site_name", "emissions_kgCO2e", "emissions_tCO2e"])
    else:
        total_kg = 0.0
        total_t = 0.0
        df_scope = pd.DataFrame(columns=["scope", "emissions_kgCO2e", "emissions_tCO2e", "percent_of_total"])
        df_scope_category = pd.DataFrame(columns=["scope", "category", "emissions_kgCO2e", "emissions_tCO2e"])
        df_top_categories = pd.DataFrame(columns=["category", "emissions_kgCO2e", "emissions_tCO2e"])
        df_sites = pd.DataFrame(columns=["site_name", "emissions_kgCO2e", "emissions_tCO2e"])

    data_quality = _build_combined_data_quality(results, line_items, all_errors, parser_diagnostics)
    confidence_score = _build_confidence_score(data_quality)

    analytics: Dict[str, Any] = {
        "average_emissions_per_row_kgco2e": round(float(df_line["emissions_kgCO2e"].mean()), 4)
        if not df_line.empty and "emissions_kgCO2e" in df_line.columns
        else 0.0
    }
    if not df_top_categories.empty:
        analytics["top_category_by_kgco2e"] = df_top_categories.iloc[0].to_dict()
        analytics["top_categories_by_kgco2e"] = df_top_categories.to_dict(orient="records")
    if not df_sites.empty:
        analytics["top_site_by_kgco2e"] = df_sites.iloc[0].to_dict()
        analytics["top_sites_by_kgco2e"] = df_sites.to_dict(orient="records")
    if not df_scope.empty:
        analytics["top_scope_by_kgco2e"] = df_scope.iloc[0].to_dict()

    plain_language_takeaways = build_plain_language_takeaways(
        total_kg=total_kg,
        total_t=total_t,
        df_scope=df_scope,
        df_top_categories=df_top_categories,
        data_quality=data_quality,
        confidence_score=confidence_score,
    )

    actionable_insights = build_actionable_insights(
        df_scope=df_scope,
        df_top_categories=df_top_categories,
        data_quality=data_quality,
        confidence_score=confidence_score,
        issue_groups=_combine_issue_groups(issue_groups),
    )

    return make_json_safe(
        {
            "totals": {"total_kgCO2e": total_kg, "total_tCO2e": total_t},
            "summary_by_scope": df_scope.to_dict(orient="records"),
            "summary_by_scope_category": df_scope_category.to_dict(orient="records"),
            "line_items": line_items,
            "analytics": analytics,
            "top_categories_by_kgco2e": df_top_categories.to_dict(orient="records"),
            "top_sites_by_kgco2e": df_sites.to_dict(orient="records"),
            "plain_language_takeaways": plain_language_takeaways,
            "actionable_insights": actionable_insights,
            "confidence_score": confidence_score,
            "data_quality": data_quality,
            "parser_diagnostics": parser_diagnostics,
            "errors": all_errors,
            "errors_preview": all_errors[:MAX_PREVIEW_ERRORS],
            "error_summary": _build_error_summary(all_errors),
            "issue_groups": _combine_issue_groups(issue_groups),
            "unmapped_categories": _merge_count_dicts(unmapped_categories, ["category", "label"]),
            "factor_transparency": factor_transparency,
            "factor_quality_summary": _merge_count_dicts(factor_quality_summary, ["factor_quality", "label"]),
            "methodology_summary": build_methodology_summary(data_quality),
        }
    )


def build_plain_language_takeaways(
    total_kg: float,
    total_t: float,
    df_scope: pd.DataFrame,
    df_top_categories: pd.DataFrame,
    data_quality: Dict[str, Any],
    confidence_score: Dict[str, Any],
) -> List[str]:
    takeaways: List[str] = []

    if total_kg > 0:
        takeaways.append(f"Total reported emissions are {round(total_t, 2)} tCO2e ({round(total_kg, 2)} kgCO2e).")
    else:
        takeaways.append("No reportable emissions were calculated from the uploaded data.")

    if not df_scope.empty:
        top_scope = df_scope.iloc[0]
        takeaways.append(
            f"{top_scope.get('scope')} is the largest contributor at {round(float(top_scope.get('percent_of_total', 0)), 2)}% of total emissions."
        )

    if not df_top_categories.empty:
        top_category = df_top_categories.iloc[0]
        takeaways.append(
            f"The highest-emitting category is {top_category.get('category')}, contributing {round(float(top_category.get('emissions_kgCO2e', 0)), 2)} kgCO2e."
        )

    takeaways.append(
        f"Data coverage is {data_quality.get('coverage_percent', 0)}%, based on extracted activity rows successfully included in the calculation."
    )

    if data_quality.get("starter_factor_rows", 0):
        takeaways.append(
            f"{data_quality.get('starter_factor_rows')} row(s) used starter/placeholder factors and should be replaced with official factor rows before final audit release."
        )

    takeaways.append(
        f"The current confidence score is {confidence_score.get('score', 0)}/100, rated {confidence_score.get('label', 'Low')}."
    )

    return takeaways[:6]


def build_actionable_insights(
    df_scope: pd.DataFrame,
    df_top_categories: pd.DataFrame,
    data_quality: Dict[str, Any],
    confidence_score: Dict[str, Any],
    issue_groups: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    insights: List[Dict[str, Any]] = []

    if data_quality.get("extracted_activity_count", data_quality.get("total_rows", 0)) == 0:
        insights.append(
            {
                "title": "Parser review required",
                "body": "No calculable activity rows were extracted. Check workbook structure, section labels, parser options, or use a mapped upload template.",
            }
        )

    if not df_scope.empty:
        top_scope = df_scope.iloc[0]
        insights.append(
            {
                "title": "Largest emission driver",
                "body": f"{top_scope.get('scope')} is currently the biggest contributor at {round(float(top_scope.get('emissions_kgCO2e', 0)), 2)} kgCO2e.",
            }
        )

    if not df_top_categories.empty:
        top_category = df_top_categories.iloc[0]
        insights.append(
            {
                "title": "Priority reduction opportunity",
                "body": f"The category with the greatest impact is {top_category.get('category')}. This is the first area to investigate for reductions and data-quality review.",
            }
        )

    if data_quality.get("coverage_percent", 0) < 90 and data_quality.get("total_rows", 0) > 0:
        insights.append(
            {
                "title": "Coverage improvement needed",
                "body": f"Coverage is currently {data_quality.get('coverage_percent', 0)}%. Resolve parser warnings, validation errors, and factor matching failures before relying on totals externally.",
            }
        )

    if data_quality.get("starter_factor_rows", 0):
        insights.append(
            {
                "title": "Replace starter factors",
                "body": f"{data_quality.get('starter_factor_rows')} row(s) used starter/placeholder factors. Replace these with official DEFRA/UK Government flat-file factor rows for audit-grade reporting.",
            }
        )

    if data_quality.get("factor_fallback_rows", 0):
        insights.append(
            {
                "title": "Review fallback matches",
                "body": f"{data_quality.get('factor_fallback_rows')} row(s) used fallback factor matching. Review these line items before client release.",
            }
        )

    if issue_groups:
        top_issue = issue_groups[0]
        insights.append(
            {
                "title": f"Top issue: {top_issue.get('title')}",
                "body": top_issue.get("guidance", "Review the issue group before issuing the report."),
            }
        )

    if confidence_score.get("score", 0) < 60:
        insights.append(
            {
                "title": "Low-confidence result",
                "body": "Input cleanup, official factor loading, and parser review should be prioritised before external reporting.",
            }
        )

    return insights[:8]


def build_methodology_summary(data_quality: Dict[str, Any]) -> Dict[str, Any]:
    assumptions = [
        "Natural gas kWh defaults to Gross CV unless Net CV or Gross CV is explicitly supplied.",
        "Miles are converted to vehicle kilometres using 1 mile = 1.609344 km.",
        "Fleet fuel litres are included by default.",
        "Business mileage is excluded by default where direct fleet fuel is present unless explicitly enabled.",
        "EV charging sub-meter data is excluded by default unless explicitly enabled, to avoid double counting against main electricity.",
    ]

    if data_quality.get("starter_factor_rows", 0):
        assumptions.append("Some rows used starter placeholder factors and require official factor replacement before final reporting.")

    return {
        "framework": "GHG Protocol aligned reporting structure using activity data multiplied by emissions factors.",
        "factor_basis": "Engine supports DEFRA/UK Government factor-year handling for 2023, 2024, and 2025. Built-in starter rows should be replaced or supplemented with official flat-file factor rows for audit use.",
        "calculation_logic": "Each valid row is mapped to a canonical category, normalised to a supported unit, matched to an emissions factor, multiplied by the activity quantity, and aggregated into CO2e outputs.",
        "assumptions": assumptions,
        "interpretation_notes": [
            "Review line-item factor transparency, parser diagnostics, excluded sections, and data-quality metrics alongside totals.",
            "Do not issue an external report if extracted activity count is zero, coverage is materially incomplete, or placeholder/fallback factor rows remain unresolved.",
        ],
        "api_version": APP_VERSION,
        "generated_at_utc": utc_now(),
        "data_quality_snapshot": data_quality,
    }


def merge_file_errors_with_result(base_result: Dict[str, Any], file_level_errors: List[Dict[str, Any]]) -> Dict[str, Any]:
    result = dict(base_result)
    combined_errors = file_level_errors + list(result.get("errors", []) or [])
    result["errors"] = combined_errors
    result["errors_preview"] = combined_errors[:MAX_PREVIEW_ERRORS]
    result["error_count"] = len(combined_errors)
    result["error_summary"] = _build_error_summary(combined_errors)
    return result


def build_api_response(
    full_result: Dict[str, Any],
    uploaded_files_info: List[Dict[str, Any]],
    input_row_count: int,
    total_upload_bytes: int,
    user: Dict[str, Any],
    parser_options: Dict[str, Any],
    request_id: str,
    started_at: float,
    privacy_consent: bool = True,
    contact_consent: bool = True,
) -> Dict[str, Any]:
    elapsed_ms = round((time.perf_counter() - started_at) * 1000, 2)
    errors = full_result.get("errors", []) or []

    response = {
        "request_id": request_id,
        "status": "completed",
        "app_version": APP_VERSION,
        "generated_at_utc": utc_now(),
        "elapsed_ms": elapsed_ms,
        "totals": full_result.get("totals", {}),
        "summary_by_scope": full_result.get("summary_by_scope", []),
        "summary_by_scope_category": full_result.get("summary_by_scope_category", []),
        "line_items": full_result.get("line_items", []),
        "uploaded_files": uploaded_files_info,
        "input_row_count": input_row_count,
        "total_upload_bytes": total_upload_bytes,
        "privacy_consent": privacy_consent,
        "contact_consent": contact_consent,
        "user": user,
        "parser_options": parser_options,
        "analytics": full_result.get("analytics", {}),
        "top_categories_by_kgco2e": full_result.get("top_categories_by_kgco2e", []),
        "top_sites_by_kgco2e": full_result.get("top_sites_by_kgco2e", []),
        "plain_language_takeaways": full_result.get("plain_language_takeaways", []),
        "actionable_insights": full_result.get("actionable_insights", []),
        "confidence_score": full_result.get("confidence_score", {}),
        "methodology_summary": full_result.get("methodology_summary", {}),
        "factor_quality_summary": full_result.get("factor_quality_summary", []),
        "error_count": len(errors),
        "errors": errors,
        "errors_preview": full_result.get("errors_preview", errors[:MAX_PREVIEW_ERRORS]),
        "error_summary": full_result.get("error_summary", []),
        "issue_groups": full_result.get("issue_groups", []),
        "data_quality": full_result.get("data_quality", {}),
        "unmapped_categories": full_result.get("unmapped_categories", []),
        "factor_transparency": full_result.get("factor_transparency", []),
        "parser_diagnostics": full_result.get("parser_diagnostics", []),
    }
    return make_json_safe(response)


# ============================================================
# LEAD LOGGING
# ============================================================

def send_lead_to_logger(user_data: Dict[str, Any], result: Dict[str, Any], uploaded_files_info: List[Dict[str, Any]]) -> None:
    url = os.getenv("LEAD_LOGGER_URL")
    if not url:
        return

    totals = result.get("totals", {}) or {}
    analytics = result.get("analytics", {}) or {}
    confidence = result.get("confidence_score", {}) or {}
    data_quality = result.get("data_quality", {}) or {}

    payload = {
        "request_id": result.get("request_id", ""),
        "full_name": user_data.get("full_name", ""),
        "email": user_data.get("email", ""),
        "company_name": user_data.get("company_name", ""),
        "phone_number": user_data.get("phone_number", ""),
        "privacy_consent": user_data.get("privacy_consent", False),
        "contact_consent": user_data.get("contact_consent", False),
        "uploaded_files_count": len(uploaded_files_info),
        "input_row_count": result.get("input_row_count", 0),
        "extracted_activity_count": data_quality.get("extracted_activity_count", 0),
        "total_kgco2e": totals.get("total_kgCO2e", 0),
        "total_tco2e": totals.get("total_tCO2e", 0),
        "top_category": (analytics.get("top_category_by_kgco2e") or {}).get("category", ""),
        "top_site": (analytics.get("top_site_by_kgco2e") or {}).get("site_name", ""),
        "confidence_score": confidence.get("score", 0),
        "confidence_label": confidence.get("label", ""),
        "error_count": result.get("error_count", 0),
        "source": "website_calculator",
        "generated_at_utc": utc_now(),
    }

    try:
        requests.post(url, json=payload, timeout=5)
    except Exception as exc:
        print(f"Lead logger failed: {type(exc).__name__}: {str(exc)}")


# ============================================================
# EXCEL REPORT BUILDERS
# ============================================================

def auto_fit_columns(ws, min_width: int = 12, max_width: int = 55) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                length = max(length, len(value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def style_report_sheet(ws, title: str = "GS Carbon Emissions Audit Report", subtitle: str = "Generated from uploaded datasets") -> None:
    ws["A1"] = title
    ws["A2"] = subtitle
    ws["A3"] = f"Generated at UTC: {utc_now()}"

    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=11, italic=True)
    ws["A3"].font = Font(size=10)

    header_fill = PatternFill(fill_type="solid", start_color="D9EAD3", end_color="D9EAD3")
    thin = Side(border_style="thin", color="DDDDDD")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    if ws.max_row >= 7:
        for cell in ws[7]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

    logo_path = os.getenv("REPORT_LOGO_PATH", "logo.png")
    if os.path.exists(logo_path):
        try:
            logo = XLImage(logo_path)
            logo.width = 140
            logo.height = 70
            ws.add_image(logo, "L1")
        except Exception:
            pass

    ws.freeze_panes = "A8"


def write_df(writer, sheet_name: str, df: pd.DataFrame) -> None:
    sheet_name = sanitize_sheet_name(sheet_name)
    if df is None or df.empty:
        pd.DataFrame([{"message": "No data available"}]).to_excel(writer, sheet_name=sheet_name, index=False, startrow=6)
    else:
        safe_df = df.copy()
        for col in safe_df.columns:
            safe_df[col] = safe_df[col].apply(lambda x: json.dumps(x, default=str) if isinstance(x, (dict, list)) else x)
        safe_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=6)


def add_bar_chart(ws, title: str, data_col: int, category_col: int, start_row: int, end_row: int, anchor: str) -> None:
    if end_row <= start_row:
        return
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "kg CO2e"
    chart.x_axis.title = "Category"
    data = Reference(ws, min_col=data_col, min_row=start_row, max_row=end_row)
    categories = Reference(ws, min_col=category_col, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, anchor)


def add_pie_chart(ws, title: str, data_col: int, category_col: int, start_row: int, end_row: int, anchor: str) -> None:
    if end_row <= start_row:
        return
    chart = PieChart()
    chart.title = title
    data = Reference(ws, min_col=data_col, min_row=start_row, max_row=end_row)
    labels = Reference(ws, min_col=category_col, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.height = 8
    chart.width = 12
    ws.add_chart(chart, anchor)


def build_overview_dataframe(result: Dict[str, Any], uploaded_files_info: List[Dict[str, Any]]) -> pd.DataFrame:
    totals = result.get("totals", {}) or {}
    confidence = result.get("confidence_score", {}) or {}
    dq = result.get("data_quality", {}) or {}

    return pd.DataFrame(
        [
            {
                "request_id": result.get("request_id", ""),
                "generated_at_utc": result.get("generated_at_utc", ""),
                "total_kgCO2e": totals.get("total_kgCO2e", 0),
                "total_tCO2e": totals.get("total_tCO2e", 0),
                "uploaded_files_count": len(uploaded_files_info),
                "input_row_count": result.get("input_row_count", 0),
                "extracted_activity_count": dq.get("extracted_activity_count", 0),
                "successful_rows": dq.get("successful_rows", 0),
                "errored_rows": dq.get("errored_rows", 0),
                "coverage_percent": dq.get("coverage_percent", 0),
                "confidence_score": confidence.get("score", 0),
                "confidence_label": confidence.get("label", ""),
                "starter_factor_rows": dq.get("starter_factor_rows", 0),
                "factor_fallback_rows": dq.get("factor_fallback_rows", 0),
                "error_count": result.get("error_count", 0),
                "app_version": result.get("app_version", APP_VERSION),
            }
        ]
    )


def write_executive_summary_sheet(writer, result: Dict[str, Any]) -> None:
    workbook = writer.book
    ws = workbook.create_sheet("Executive Summary")

    style_report_sheet(ws, "GS Carbon Emissions Audit Report", "Executive summary")

    totals = result.get("totals", {}) or {}
    confidence = result.get("confidence_score", {}) or {}
    dq = result.get("data_quality", {}) or {}
    takeaways = result.get("plain_language_takeaways", []) or []
    insights = result.get("actionable_insights", []) or []
    scope_rows = result.get("summary_by_scope", []) or []
    top_categories = result.get("top_categories_by_kgco2e", []) or []
    issue_groups = result.get("issue_groups", []) or []

    metrics = [
        ("Total emissions (tCO2e)", totals.get("total_tCO2e", 0)),
        ("Total emissions (kgCO2e)", totals.get("total_kgCO2e", 0)),
        ("Confidence score", confidence.get("score", 0)),
        ("Confidence label", confidence.get("label", "")),
        ("Coverage percent", dq.get("coverage_percent", 0)),
        ("Extracted activity count", dq.get("extracted_activity_count", 0)),
        ("Successful rows", dq.get("successful_rows", 0)),
        ("Error count", result.get("error_count", 0)),
        ("Starter factor rows", dq.get("starter_factor_rows", 0)),
        ("Fallback factor rows", dq.get("factor_fallback_rows", 0)),
    ]

    ws["A6"] = "Metric"
    ws["B6"] = "Value"
    for i, (metric, value) in enumerate(metrics, start=7):
        ws[f"A{i}"] = metric
        ws[f"B{i}"] = value

    ws["D6"] = "Plain-language takeaways"
    for i, item in enumerate(takeaways or ["No takeaways available."], start=7):
        ws[f"D{i}"] = item

    ws["G6"] = "Actionable insights"
    ws["H6"] = "Details"
    for i, item in enumerate(insights or [{"title": "No insights available.", "body": ""}], start=7):
        ws[f"G{i}"] = item.get("title", "")
        ws[f"H{i}"] = item.get("body", "")

    start = 20
    ws[f"A{start}"] = "Scope"
    ws[f"B{start}"] = "kg CO2e"
    ws[f"C{start}"] = "% of total"
    for i, row in enumerate(scope_rows, start=start + 1):
        ws[f"A{i}"] = row.get("scope")
        ws[f"B{i}"] = row.get("emissions_kgCO2e")
        ws[f"C{i}"] = row.get("percent_of_total")

    ws[f"E{start}"] = "Top category"
    ws[f"F{start}"] = "kg CO2e"
    for i, row in enumerate(top_categories, start=start + 1):
        ws[f"E{i}"] = row.get("category")
        ws[f"F{i}"] = row.get("emissions_kgCO2e")

    ws[f"H{start}"] = "Issue"
    ws[f"I{start}"] = "Severity"
    ws[f"J{start}"] = "Count"
    ws[f"K{start}"] = "Guidance"
    for i, row in enumerate(issue_groups, start=start + 1):
        ws[f"H{i}"] = row.get("title")
        ws[f"I{i}"] = row.get("severity")
        ws[f"J{i}"] = row.get("count")
        ws[f"K{i}"] = row.get("guidance")

    auto_fit_columns(ws)


def write_charts_sheet(writer, result: Dict[str, Any]) -> None:
    workbook = writer.book
    ws = workbook.create_sheet("Charts")

    style_report_sheet(ws, "GS Carbon Emissions Audit Report", "Visual summary")

    scope_rows = result.get("summary_by_scope", []) or []
    cat_rows = result.get("summary_by_scope_category", []) or []

    ws["A6"] = "Scope"
    ws["B6"] = "kg CO2e"
    for i, row in enumerate(scope_rows, start=7):
        ws[f"A{i}"] = row.get("scope")
        ws[f"B{i}"] = row.get("emissions_kgCO2e")

    ws["D6"] = "Scope - Category"
    ws["E6"] = "kg CO2e"
    for i, row in enumerate(cat_rows, start=7):
        ws[f"D{i}"] = f"{row.get('scope')} - {row.get('category')}"
        ws[f"E{i}"] = row.get("emissions_kgCO2e")

    ws["J6"] = "Scope"
    ws["K6"] = "kg CO2e"
    for i, row in enumerate(scope_rows, start=7):
        ws[f"J{i}"] = row.get("scope")
        ws[f"K{i}"] = row.get("emissions_kgCO2e")

    ws["N6"] = "Scope - Category"
    ws["O6"] = "kg CO2e"
    for i, row in enumerate(cat_rows, start=7):
        ws[f"N{i}"] = f"{row.get('scope')} - {row.get('category')}"
        ws[f"O{i}"] = row.get("emissions_kgCO2e")

    if scope_rows:
        add_bar_chart(ws, "Emissions by Scope", 2, 1, 6, 6 + len(scope_rows), "G6")
        add_pie_chart(ws, "Scope Share", 11, 10, 6, 6 + len(scope_rows), "Q6")

    if cat_rows:
        add_bar_chart(ws, "Emissions by Scope and Category", 5, 4, 6, 6 + len(cat_rows), "G24")
        add_pie_chart(ws, "Category Share", 15, 14, 6, 6 + len(cat_rows), "Q24")

    auto_fit_columns(ws)


def build_excel_report(result: Dict[str, Any], uploaded_files_info: List[Dict[str, Any]]) -> io.BytesIO:
    output = io.BytesIO()

    dfs = {
        "Overview": build_overview_dataframe(result, uploaded_files_info),
        "Summary by Scope": pd.DataFrame(result.get("summary_by_scope", [])),
        "Summary by Category": pd.DataFrame(result.get("summary_by_scope_category", [])),
        "Top Categories": pd.DataFrame(result.get("top_categories_by_kgco2e", [])),
        "Top Sites": pd.DataFrame(result.get("top_sites_by_kgco2e", [])),
        "Takeaways": pd.DataFrame([{"takeaway": x} for x in result.get("plain_language_takeaways", [])]),
        "Insights": pd.DataFrame(result.get("actionable_insights", [])),
        "Confidence": pd.DataFrame([result.get("confidence_score", {})]),
        "Data Quality": pd.DataFrame([result.get("data_quality", {})]),
        "Parser Diagnostics": pd.DataFrame(result.get("parser_diagnostics", [])),
        "Parser Options": pd.DataFrame([result.get("parser_options", {})]),
        "Issue Groups": pd.DataFrame(result.get("issue_groups", [])),
        "Errors": pd.DataFrame(result.get("errors", [])),
        "Error Summary": pd.DataFrame(result.get("error_summary", [])),
        "Uploaded Files": pd.DataFrame(uploaded_files_info),
        "Line Items": pd.DataFrame(result.get("line_items", [])),
        "Factor Transparency": pd.DataFrame(result.get("factor_transparency", [])),
        "Factor Quality": pd.DataFrame(result.get("factor_quality_summary", [])),
        "Unmapped Categories": pd.DataFrame(result.get("unmapped_categories", [])),
        "Methodology": pd.DataFrame([result.get("methodology_summary", {})]),
        "User Details": pd.DataFrame([result.get("user", {})]),
    }

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            write_df(writer, sheet_name, df)

        for sheet_name in list(writer.book.sheetnames):
            ws = writer.book[sheet_name]
            style_report_sheet(ws)
            auto_fit_columns(ws)

        write_executive_summary_sheet(writer, result)
        write_charts_sheet(writer, result)

    output.seek(0)
    return output


# ============================================================
# API ROUTES
# ============================================================

@app.post("/calculate")
def calculate(payload: Dict[str, Any]) -> Dict[str, Any]:
    request_id = new_request_id()
    started_at = time.perf_counter()

    try:
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="Payload must be a JSON object.")

        activities = payload.get("activities", [])
        parser_options = build_parser_options(payload.get("parser_options", {}))
        factor_rows = payload.get("factor_rows")

        if not isinstance(activities, list):
            raise HTTPException(status_code=400, detail="activities must be a list.")

        if factor_rows is not None and not isinstance(factor_rows, list):
            raise HTTPException(status_code=400, detail="factor_rows must be a list when provided.")

        engine_request: Dict[str, Any] = {
            "activities": activities,
            "parser_options": parser_options,
        }
        if factor_rows:
            engine_request["factor_rows"] = factor_rows

        result = run_emissions_engine(engine_request)
        result["request_id"] = request_id
        result["app_version"] = APP_VERSION
        result["generated_at_utc"] = utc_now()
        result["elapsed_ms"] = round((time.perf_counter() - started_at) * 1000, 2)
        return make_json_safe(result)

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {str(exc)}")


@app.post("/upload-calculate")
async def upload_calculate(
    files: Annotated[List[UploadFile], File(description="Upload one or more CSV/XLS/XLSX activity datasets")],
    privacy_consent: Annotated[str, Form(...)],
    contact_consent: Annotated[str, Form(...)],
    full_name: Annotated[str, Form(...)],
    email: Annotated[str, Form(...)],
    company_name: Annotated[str, Form(...)],
    phone_number: Annotated[str, Form(...)],
    include_fleet_fuel: Annotated[str, Form("true")],
    include_business_mileage: Annotated[str, Form("false")],
    include_business_mileage_when_fuel_present: Annotated[str, Form("false")],
    include_ev_charging_submeter: Annotated[str, Form("false")],
    fleet_fuel_type: Annotated[str, Form("Diesel")],
    electricity_country: Annotated[str, Form("Electricity: UK")],
    gas_unit: Annotated[str, Form("kWh (Gross CV)")],
    factor_rows_json: Annotated[Optional[str], Form(None)] = None,
) -> Dict[str, Any]:
    request_id = new_request_id()
    started_at = time.perf_counter()

    try:
        validate_user_inputs(privacy_consent, contact_consent, full_name, email, company_name, phone_number)

        user = build_user_object(full_name, email, company_name, phone_number)
        parser_options = build_parser_options(
            {
                "include_fleet_fuel": include_fleet_fuel,
                "include_business_mileage": include_business_mileage,
                "include_business_mileage_when_fuel_present": include_business_mileage_when_fuel_present,
                "include_ev_charging_submeter": include_ev_charging_submeter,
                "fleet_fuel_type": fleet_fuel_type,
                "electricity_country": electricity_country,
                "gas_unit": gas_unit,
            }
        )
        factor_rows = parse_factor_rows_json(factor_rows_json)

        uploaded_sources, file_level_errors, input_row_count, total_upload_bytes = prepare_uploaded_sources(files)

        uploaded_files_info = [
            {
                "filename": src.get("source_file"),
                "sheet_name": src.get("sheet_name"),
                "source_type": src.get("source_type"),
                "rows": src.get("rows", 0),
                "columns": src.get("columns", []),
                "size_bytes": src.get("size_bytes", 0),
            }
            for src in uploaded_sources
        ]

        full_result = run_engine_for_uploaded_sources(uploaded_sources, parser_options, factor_rows=factor_rows)
        full_result = merge_file_errors_with_result(full_result, file_level_errors)

        api_result = build_api_response(
            full_result=full_result,
            uploaded_files_info=uploaded_files_info,
            input_row_count=input_row_count,
            total_upload_bytes=total_upload_bytes,
            user=user,
            parser_options=parser_options,
            request_id=request_id,
            started_at=started_at,
            privacy_consent=True,
            contact_consent=True,
        )

        send_lead_to_logger(
            {
                **user,
                "privacy_consent": True,
                "contact_consent": True,
            },
            api_result,
            uploaded_files_info,
        )

        return api_result

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {str(exc)}")


@app.post("/upload-calculate-download")
async def upload_calculate_download(
    files: Annotated[List[UploadFile], File(description="Upload one or more CSV/XLS/XLSX activity datasets")],
    privacy_consent: Annotated[str, Form(...)],
    contact_consent: Annotated[str, Form(...)],
    full_name: Annotated[str, Form(...)],
    email: Annotated[str, Form(...)],
    company_name: Annotated[str, Form(...)],
    phone_number: Annotated[str, Form(...)],
    include_fleet_fuel: Annotated[str, Form("true")],
    include_business_mileage: Annotated[str, Form("false")],
    include_business_mileage_when_fuel_present: Annotated[str, Form("false")],
    include_ev_charging_submeter: Annotated[str, Form("false")],
    fleet_fuel_type: Annotated[str, Form("Diesel")],
    electricity_country: Annotated[str, Form("Electricity: UK")],
    gas_unit: Annotated[str, Form("kWh (Gross CV)")],
    factor_rows_json: Annotated[Optional[str], Form(None)] = None,
):
    request_id = new_request_id()
    started_at = time.perf_counter()

    try:
        validate_user_inputs(privacy_consent, contact_consent, full_name, email, company_name, phone_number)

        user = build_user_object(full_name, email, company_name, phone_number)
        parser_options = build_parser_options(
            {
                "include_fleet_fuel": include_fleet_fuel,
                "include_business_mileage": include_business_mileage,
                "include_business_mileage_when_fuel_present": include_business_mileage_when_fuel_present,
                "include_ev_charging_submeter": include_ev_charging_submeter,
                "fleet_fuel_type": fleet_fuel_type,
                "electricity_country": electricity_country,
                "gas_unit": gas_unit,
            }
        )
        factor_rows = parse_factor_rows_json(factor_rows_json)

        uploaded_sources, file_level_errors, input_row_count, total_upload_bytes = prepare_uploaded_sources(files)

        uploaded_files_info = [
            {
                "filename": src.get("source_file"),
                "sheet_name": src.get("sheet_name"),
                "source_type": src.get("source_type"),
                "rows": src.get("rows", 0),
                "columns": src.get("columns", []),
                "size_bytes": src.get("size_bytes", 0),
            }
            for src in uploaded_sources
        ]

        full_result = run_engine_for_uploaded_sources(uploaded_sources, parser_options, factor_rows=factor_rows)
        full_result = merge_file_errors_with_result(full_result, file_level_errors)

        api_result = build_api_response(
            full_result=full_result,
            uploaded_files_info=uploaded_files_info,
            input_row_count=input_row_count,
            total_upload_bytes=total_upload_bytes,
            user=user,
            parser_options=parser_options,
            request_id=request_id,
            started_at=started_at,
            privacy_consent=True,
            contact_consent=True,
        )

        output = build_excel_report(api_result, uploaded_files_info)

        filename = REPORT_FILENAME
        if request_id:
            filename = REPORT_FILENAME.replace(".xlsx", f"_{request_id}.xlsx")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {str(exc)}")
